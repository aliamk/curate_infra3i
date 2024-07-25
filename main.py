import streamlit as st
import pandas as pd
import tempfile
import os
from openpyxl import load_workbook
from datetime import datetime
import time
import pytz
import re

# Function to autofit columns
def autofit_columns(file_path):
    workbook = load_workbook(file_path)
    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(file_path)

# Function to safely convert values to float
def safe_float_conversion(value):
    if isinstance(value, str):
        value = value.replace(',', '').strip()
    try:
        return float(value)
    except ValueError:
        return 0

# Function to process 'Transaction' tab specific replacements
def process_transaction_sheet(transaction_df):
    # Replace specific words in the 'Transaction Status' column
    transaction_df['Transaction Status'] = transaction_df['Transaction Status'].replace({
        "Binding Bids": "Preparation",
        "Expressions of Interest": "Preparation",
        "Indicative Bids": "Preparation",
        "Pre-Launch": "Preparation",
        "Pre-Qualified Proponents": "Preparation",
        "RFP Returned": "Preparation",
        "RFQ returned": "Preparation",
        "Shortlisted Proponents": "Preparation",
        "Transaction Launch": "Preparation",
        "On Hold": "Preparation",
        "Preferred Proponent": "Financing",
        "No Private Financing": ""
    }, regex=True)
    
    # Replace specific words in the 'Transaction Type' column
    transaction_df['Transaction Type'] = transaction_df['Transaction Type'].apply(replace_transaction_type)
    
    # Replace specific words in the 'Region - Country' column
    transaction_df['Region - Country'] = transaction_df['Region - Country'].apply(replace_region_country)
    
    # Replace specific words in the 'Contract' column
    transaction_df['Contract'] = transaction_df['Contract'].apply(replace_contract)
    
    # Clean the 'Transaction Name' column
    transaction_df['Transaction Name'] = transaction_df['Transaction Name'].str.replace(' and ', ' & ').str.strip()
    transaction_df['Transaction Name'] = transaction_df['Transaction Name'].apply(lambda x: re.sub(r'\s{2,}', ' ', x))

    # Set 'Asset Class' column to 'Infrastructure'
    transaction_df['Transaction Asset Class'] = 'Infrastructure'


def replace_tranche_tertiary_type(tranches_df):
    replacements = {
        'Capex Facility': '',
        'Change-in-Law Facility': '',
        'Equity Bridge Loan': '',
        'Export Credit': 'Export Credit Facility',
        'Government Grant': '',
        'Government Loan': 'State Loan',
        'Islamic Financing': 'Term Loan',
        'Multilateral': 'Multilateral Loan',
        'Other': '',
        'Standby/Contigency Facility': 'Standby Facility'
    }
    tranches_df['Tranche Tertiary Type'] = tranches_df['Tranche Tertiary Type'].replace(replacements)
    return tranches_df

def replace_transaction_type(type):
    replacements = {
        'Additional Financing': 'Additional Financing',
        'Greenfield': 'Primary Financing',
        'M&A': 'Acquisition',
        'Nationalisation': '',
        'Privatisation': 'Privatisation',
        'Privatisation,M&A': 'Privatisation',
        'Public Offering': '',
        'Refinancing': 'Refinancing',
        'Take Private': ''
    }
    return replacements.get(type, type)

def replace_region_country(region):
    replacements = {
        'AFGHANISTAN': 'Afghanistan',
        'ALBANIA': 'Albania',
        'ALGERIA': 'Algeria',
        'ANDORRA': 'Andorra',
        'ANGOLA': 'Angola',
        'ARGENTINA': 'Argentina',
        'ARMENIA': 'Armenia',
        'ARUBA': 'Aruba',
        'AUSTRALIA': 'Australia',
        'AUSTRIA': 'Austria',
        'AZERBAIJAN': 'Azerbaijan',
        'BAHAMAS': 'Bahamas',
        'BAHRAIN': 'Bahrain',
        'BANGLADESH': 'Bangladesh',
        'BARBADOS': 'Barbados',
        'BELARUS': 'Belarus',
        'BELGIUM': 'Belgium',
        'BENIN': 'Benin',
        'BERMUDA': 'Bermuda',
        'BOLIVIA': 'Bolivia',
        'BOSNIA': 'Bosnia & Herzegovina',
        'BOTSWANA': 'Botswana',
        'BRAZIL': 'Brazil',
        'BRUNEI': 'Brunei',
        'BULGARIA': 'Bulgaria',
        'BURKINA FASO': 'Burkina Faso',
        'BURUNDI': 'Burundi',
        'CAMBODIA': 'Cambodia',
        'CAMEROON': 'Cameroon',
        'CANADA': 'Canada',
        'CAPE VERDE': 'Cape Verde',
        'CAYMAN ISLANDS': 'Cayman Islands',
        'CHAD': 'Chad',
        'CHILE': 'Chile',
        'CHINA': 'China',
        'COLOMBIA': 'Colombia',
        'CONGO - REPUBLIC OF THE': 'Republic of the Congo',
        'COSTA RICA': 'Costa Rica',
        'CROATIA': 'Croatia',
        'CURACAO': 'Curaçao',
        'CYPRUS': 'Cyprus',
        'CZECH REPUBLIC': 'Czech Republic',
        'DENMARK': 'Denmark',
        'DJIBOUTI': 'Djibouti',
        'DOMINICAN REPUBLIC': 'Dominican Republic',
        'DR CONGO': 'Democratic Republic of Congo',
        'EAST TIMOR': 'Timor-Leste',
        'ECUADOR': 'Ecuador',
        'EGYPT': 'Egypt',
        'EL SALVADOR': 'El Salvador',
        'ESTONIA': 'Estonia',
        'ETHIOPIA': 'Ethiopia',
        'FINLAND': 'Finland',
        'FRANCE': 'France',
        'FRENCH GUIANA': 'French Guiana',
        'FRENCH POLYNESIA': 'French Polynesia',
        'GABON': 'Gabon',
        'GAMBIA': 'Gambia',
        'GEORGIA': 'Georgia',
        'GERMANY': 'Germany',
        'GHANA': 'Ghana',
        'GIBRALTAR': 'Gibraltar',
        'GREECE': 'Greece',
        'GUATEMALA': 'Guatemala',
        'GUINEA': 'Guinea',
        'GUYANA': 'Guyana',
        'HONDURAS': 'Honduras',
        'HONG KONG (CHINA)': 'Hong Kong',
        'HUNGARY': 'Hungary',
        'ICELAND': 'Iceland',
        'INDIA': 'India',
        'INDONESIA': 'Indonesia',
        'IRAQ': 'Iraq',
        'IRELAND': 'Ireland',
        'ISRAEL': 'Israel',
        'ITALY': 'Italy',
        'IVORY COAST': 'Ivory Coast',
        'JAMAICA': 'Jamaica',
        'JAPAN': 'Japan',
        'JORDAN': 'Jordan',
        'KAZAKHSTAN': 'Kazakhstan',
        'KENYA': 'Kenya',
        'KOSOVO': 'Kosovo',
        'KUWAIT': 'Kuwait',
        'KYRGYZSTAN': 'Kyrgyzstan',
        'LAOS': 'Laos',
        'LATVIA': 'Latvia',
        'LIBERIA': 'Liberia',
        'LIBYA': 'Libya',
        'LITHUANIA': 'Lithuania',
        'LUXEMBOURG': 'Luxembourg',
        'MADAGASCAR': 'Madagascar',
        'MALAWI': 'Malawi',
        'MALAYSIA': 'Malaysia',
        'MALDIVES': 'Maldives',
        'MALI': 'Mali',
        'MAURITIUS': 'Mauritius',
        'MEXICO': 'Mexico',
        'MOLDOVA': 'Moldova',
        'MONACO': 'Monaco',
        'MONGOLIA': 'Mongolia',
        'MONTENEGRO': 'Montenegro',
        'MONTSERRAT': 'Montserrat',
        'MOROCCO': 'Morocco',
        'MOZAMBIQUE': 'Mozambique',
        'MYANMAR': 'Myanmar',
        'NAMIBIA': 'Namibia',
        'NEPAL': 'Nepal',
        'NETHERLANDS': 'Netherlands',
        'NETHERLANDS ANTILLES': '',
        'NEW ZEALAND': 'New Zealand',
        'NICARAGUA': 'Nicaragua',
        'NIGER': 'Niger',
        'NIGERIA': 'Nigeria',
        'NORTH MACEDONIA': 'North Macedonia',
        'NORWAY': 'Norway',
        'OMAN': 'Oman',
        'PAKISTAN': 'Pakistan',
        'PALESTINE': 'Palestine',
        'PANAMA': 'Panama',
        'PAPUA NEW GUINEA': 'Papua New Guinea',
        'PARAGUAY': 'Paraguay',
        'PERU': 'Peru',
        'PHILIPPINES': 'Philippines',
        'POLAND': 'Poland',
        'PORTUGAL': 'Portugal',
        'QATAR': 'Qatar',
        'REUNION': 'Reunion',
        'ROMANIA': 'Romania',
        'RUSSIA': 'Russia',
        'RWANDA': 'Rwanda',
        'SAUDI ARABIA': 'Saudi Arabia',
        'SENEGAL': 'Senegal',
        'SERBIA': 'Serbia',
        'SEYCHELLES': 'Seychelles',
        'SINGAPORE': 'Singapore',
        'SLOVAKIA': 'Slovakia',
        'SLOVENIA': 'Slovenia',
        'SOUTH AFRICA': 'South Africa',
        'SOUTH KOREA': 'South Korea',
        'SPAIN': 'Spain',
        'SRI LANKA': 'Sri Lanka',
        'SWEDEN': 'Sweden',
        'SWITZERLAND': 'Switzerland',
        'SYRIA': 'Syria',
        'TAIWAN (CHINA)': 'Taiwan',
        'TAJIKISTAN': 'Tajikistan',
        'TANZANIA': 'Tanzania',
        'THAILAND': 'Thailand',
        'TOGO': 'Togo',
        'TRINIDAD & TOBAGO': 'Trinidad and Tobago',
        'TUNISIA': 'Tunisia',
        'TURKEY': 'Turkey',
        'UGANDA': 'Uganda',
        'UKRAINE': 'Ukraine',
        'UNITED ARAB EMIRATES': 'United Arab Emirates',
        'UNITED KINGDOM': 'United Kingdom',
        'URUGUAY': 'Uruguay',
        'USA': 'United States',
        'UZBEKISTAN': 'Uzbekistan',
        'VIETNAM': 'Vietnam',
        'VIRGIN ISLANDS (US)': 'US Virgin Islands',
        'ZAMBIA': 'Zambia',
        'ZIMBABWE': 'Zimbabwe'
    }
    return replacements.get(region, region)

def replace_event_type(eventtype):
    replacements = {
        'Binding Bids': '',
        'Cancelled': 'Cancelled',
        'Expressions of Interest': 'Expression of Interest',
        'Financial Close': 'Financial Close',
        'Indicative Bids': '',
        'No Private Financing': '',
        'On Hold': '',
        'Preferred Proponent': 'Preferred Bidder',
        'Pre-Launch': '',
        'Pre-Qualified Proponents': '',
        'RFP Returned': 'Request for Proposals',
        'RFQ returned': 'Request for Qualifications',
        'Shortlisted Proponents': 'Shortlist',
        'Transaction Launch': 'Announced',
    }
    return replacements.get(eventtype, eventtype)

def replace_contract(contract):
    replacements = {
        'DBFOM': 'DBFOM',
        'DBFM': 'DBFM',
        'DBFO': 'DBFO',
        'DBF': 'DBF',
        'BF': '',
        'BFOM': '',
        'DBOM': '',
        'BFO': '',
        'BO': '',
        'OM': '',
        'DBO': '',
        'DB': '',
        'FOM': '',
        'BOM': '',
        'DFOM': '',
        'DBM': '',
        'BM': '',
        'DOM': '',
        'DO': '',
        'DFO': '',
        'O': '',
    }
    return replacements.get(contract, contract)

# Dictionary for Any Level Sectors replacements
sector_replacements = {
    'Accommodation': 'Social Infrastructure',
    'Airports': 'Transport, Airport',
    'Battery Storage': 'Renewable Energy, Energy Storage',
    'Biofuels': 'Renewable Energy, Biofuels/Biomass',
    'Biogas': 'Renewable Energy, Biofuels/Biomass',
    'Biomass': 'Renewable Energy, Biofuels/Biomass',
    'Bridges and Tunnels': 'Transport',
    'Broadband': 'Digital Infrastructure, Internet',
    'Car Parks': 'Transport, Car Park',
    'Carbon Capture': 'Renewable Energy, Carbon Capture & Storage',
    'Coal fired': 'Conventional Energy, Coal-Fired Power',
    'Co-generation': 'Conventional Energy, Cogeneration Power',
    'Courthouses': 'Social Infrastructure, Justice',
    'Data Centre': 'Digital Infrastructure, Data Centre',
    'Defence': 'Social Infrastructure',
    'Desalination': 'Water, Desalination',
    'District Heating & Cooling': 'Social Infrastructure, Heat Network',
    'Education': 'Social Infrastructure, Education',
    'Electricity Distribution': 'Conventional Energy, Transmission',
    'Electricity Smart Meter': 'Conventional Energy, Transmission',
    'Electricity Transmission': 'Conventional Energy, Transmission',
    'Energy from waste': 'Renewable Energy, Waste to Energy',
    'Energy Other': 'Conventional Energy',
    'Environment': 'Renewable Energy',
    'EV Infrastructure': 'Renewable Energy, EV Charging',
    'Exploration & Production': 'Oil & Gas, Upstream',
    'Ferries': 'Transport, Waterway',
    'Fibre Optic': 'Digital Infrastructure, Internet',
    'Floating Solar PV': 'Renewable Energy, Solar (Floating PV)',
    'Gas Distribution': 'Oil & Gas, Downstream',
    'Gas fired': 'Conventional Energy, Gas-Fired Power',
    'Gas Pipeline': 'Oil & Gas, Midstream',
    'Gas Smart Meter': 'Conventional Energy',
    'Geothermal': 'Renewable Energy, Geothermal',
    'Healthcare': 'Social Infrastructure, Healthcare',
    'High-speed Rail': 'Transport, Heavy Rail',
    'Hydroelectric': 'Renewable Energy, Hydro',
    'Hydrogen': 'Renewable Energy, Hydrogen',
    'IWPP': 'Conventional Energy',
    'Leisure': 'Social Infrastructure, Leisure',
    'LNG export terminal': 'Oil & Gas, LNG',
    'Microgrids': 'Conventional Energy, Transmission',
    'Mining': 'Mining',
    'Nuclear': 'Conventional Energy, Nuclear Power',
    'Offshore wind': 'Renewable Energy, Wind (Offshore)',
    'Oil & Gas Storage': 'Oil & Gas, Midstream',
    'Oil & gas transportation': 'Oil & Gas, Midstream',
    'Oil fired': 'Conventional Energy, Oil-Fired Power',
    'Oil Pipeline': 'Oil & Gas, Midstream',
    'Onshore wind': 'Renewable Energy, Wind (Onshore)',
    'Petrochemical plants': 'Oil & Gas, Petrochemical',
    'Police Facilities': 'Social Infrastructure, Justice',
    'Ports': 'Transport, Port',
    'Power Other': 'Conventional Energy',
    'Prisons': 'Social Infrastructure, Justice',
    'Rail': 'Transport, Heavy Rail',
    'Refineries': 'Oil & Gas',
    'Renewables Other': 'Renewable Energy',
    'Renewables': 'Renewable Energy',
    'Roads': 'Transport, Road',
    'Rolling Stock': 'Transport, Heavy Rail',
    'Social Housing': 'Social Infrastructure, Social Housing',
    'Social Infrastructure Other': 'Social Infrastructure',
    'Solar CSP': 'Renewable Energy, Solar (Thermal)',
    'Solar PV': 'Renewable Energy, Solar (Land-Based Solar)',
    'Subsea Cable': 'Digital Infrastructure',
    'Telecommunications Other': 'Digital Infrastructure',
    'Tidal': 'Renewable Energy, Marine',
    'Transport Other': 'Transport',
    'Urban Rail Transit': 'Transport, Light Transport',
    'Waste': 'Waste',
    'Water': 'Water',
    'Wireless Transmission': 'Digital Infrastructure'
}

# Function to replace sector names based on the dictionary
def replace_sector_names(sector_string):
    sectors = sector_string.split(', ')
    replaced_sectors = []
    for sector in sectors:
        replaced_sectors.append(sector_replacements.get(sector, sector))
        return ', '.join(replaced_sectors)

def clean_company_names(tranche_roles_any_df):
    def clean_company_name(name):
        # a) Delete content within parenthesis and delete parenthesis
        name = re.sub(r'\s*\(.*?\)\s*', '', name)
        # b) Delete all trailing spaces
        name = name.strip()
        # c) Delete two or more spaces in between words
        name = re.sub(r'\s{2,}', ' ', name)
        return name
    
    tranche_roles_any_df['Company'] = tranche_roles_any_df['Company'].apply(clean_company_name)
    return tranche_roles_any_df

# Function to create a destination file with specific tabs and populate 'Transaction', 'Events', 'Bidders_Any', and 'Tranches' tabs
def create_destination_file(source_path, start_time):
    try:
        # Read the Excel file
        df = pd.read_excel(source_path, parse_dates=True)
        
        # Print the column names
        st.write("Column names in the uploaded file:")
        st.write(df.columns.tolist())
        
        # Create a dictionary to hold DataFrames for each tab
        tabs = {
            "Transaction": pd.DataFrame(columns=[
                "Transaction Upload ID", "Transaction Name", "Transaction Asset Class", 
                "Transaction Status", "Finance Type", "Transaction Type", "Unknown Asset", 
                "Underlying Asset Configuration", "Transaction Local Currency", 
                "Transaction Value (Local Currency)", "Transaction Debt (Local Currency)", 
                "Transaction Equity (Local Currency)", "Debt/Equity Ratio", 
                "Underlying Number of Assets", "Region - Country", "Region - State", 
                "Region - City", "Any Level Sectors", "PPP", "Concession Period", 
                "Contract", "SPV", "Active"
            ]),
            "Underlying_Asset": pd.DataFrame(columns=[
                "Transaction Upload ID", "Asset Upload ID"
            ]),
            "Events": pd.DataFrame(columns=[
                "Transaction Upload ID", "Event Date", "Event Type"
            ]),
            "Bidders_Any": pd.DataFrame(columns=[
                "Transaction Upload ID", "Role Type", "Role Subtype", "Company", 
                "Fund", "Bidder Status", "Client Counterparty", "Client Company Name", "Fund Name"
            ]),
            "Tranches": pd.DataFrame(columns=[
                "Transaction Upload ID", "Tranche Upload ID", "Tranche Primary Type", 
                "Tranche Secondary Type", "Tranche Tertiary Type", "Value", 
                "Maturity Start Date", "Maturity End Date", "Tenor", 
                "Tranche ESG Type", "Helper_Tranche Value USD m", "Helper_Transaction Value USD m", "Helper_Transaction Value LC",
                "Helper_Tranche Value USD m as % of Helper_Transaction Value USD m"
            ]),
            "Tranche_Pricings": pd.DataFrame(columns=[
                "Tranche Upload ID", "Tranche Benchmark", "Basis Point From", "Basis Point To", 
                "Period From", "Period To", "Period Duration", "Comment"
            ]),
            "Tranche_Roles_Any": pd.DataFrame(columns=[
                "Transaction Upload ID", "Tranche Upload ID", "Tranche Role Type", 
                "Company", "Fund", "Value", "Percentage", "Comment"
            ])
        }
        
        # Helper function to copy column data if the column exists in the source file
        def copy_column(src_df, src_col, dest_df, dest_col):
            if src_col in src_df.columns:
                dest_df[dest_col] = src_df[src_col]
                st.write(f"<small>Copied data from '{src_col}' to '{dest_col}'</small>", unsafe_allow_html=True)
            else:
                st.warning(f"'{src_col}' column not found in the source file.")

        # Function to handle date columns
        def format_dates(df, columns):
            for column in columns:
                if column in df.columns:
                    df[column] = pd.to_datetime(df[column], errors='coerce').dt.date

        # Format date columns in 'Events' tab
        format_dates(df, [
            "Current status date", "Financial close", "Transaction Launch", 
            "RFP returned", "Preferred Proponents", "Expressions of Interest", 
            "RFQ returned", "Shortlisted proponents"
        ])
        
        # Format date columns in 'Tranches' tab
        format_dates(df, ["Maturity Start Date", "Maturity End Date"])

        # Populating 'Transaction' tab
        copy_column(df, 'Transaction Upload ID', tabs["Transaction"], 'Transaction Upload ID')
        copy_column(df, 'Transaction Name', tabs["Transaction"], 'Transaction Name')
        copy_column(df, 'Current status', tabs["Transaction"], 'Transaction Status')
        copy_column(df, 'Type', tabs["Transaction"], 'Transaction Type')
        copy_column(df, 'Transaction size (m)', tabs["Transaction"], 'Transaction Local Currency')
        copy_column(df, 'Transaction size (m)', tabs["Transaction"], 'Transaction Value (Local Currency)')
        copy_column(df, 'Geography', tabs["Transaction"], 'Region - Country')
        copy_column(df, 'PPP', tabs["Transaction"], 'PPP')
        copy_column(df, 'Duration', tabs["Transaction"], 'Concession Period')
        copy_column(df, 'Delivery Model', tabs["Transaction"], 'Contract')
        copy_column(df, 'SPV', tabs["Transaction"], 'SPV')
        if 'Transaction Name' in df.columns:
            tabs["Transaction"]['Active'] = 'TRUE'
        
        # Add the new column 'Helper_Any Level Sectors' after 'Active' column
        tabs["Transaction"].insert(
            tabs["Transaction"].columns.get_loc("Active") + 1, 
            "Helper_Any Level Sectors", 
            None
        )

        # Function to replace sector names based on the dictionary
        def replace_sector_names(sector_string):
            sectors = sector_string.split(', ')
            replaced_sectors = []
            for sector in sectors:
                replacement = sector_replacements.get(sector, sector)
                if replacement:
                    replaced_sectors.append(replacement)
            return ', '.join(replaced_sectors)

        # Function to concatenate 'Sector' and 'Sub-Sector' with a comma and replace sector names
        def concatenate_and_replace_sectors(sector, subsector):
            concatenated = concatenate_sectors_and_subsectors(sector, subsector)
            return replace_sector_names(concatenated)

        # Function to concatenate 'Sector' and 'Sub-Sector' with a comma
        def concatenate_sectors_and_subsectors(sector, subsector):
            if pd.isna(sector) and pd.isna(subsector):
                return ""
            elif pd.isna(sector):
                return subsector
            elif pd.isna(subsector):
                return sector
            else:
                return f"{sector}, {subsector}"

        # Populate 'Helper_Any Level Sectors' column
        if 'Sector' in df.columns or 'Sub-Sector' in df.columns:
            tabs["Transaction"]["Helper_Any Level Sectors"] = df.apply(
                lambda row: concatenate_sectors_and_subsectors(row.get("Sector", ""), row.get("Sub-Sector", "")),
                axis=1
            )
            st.write("<small>Copied data for 'Helper_Any Level Sectors' from 'Sector' and 'Sub-Sector'</small>", unsafe_allow_html=True)
        else:
            st.warning("'Sector' and 'Sub-Sector' columns not found in the source file.")

        # Populate 'Any Level Sectors' column based on 'Helper_Any Level Sectors'
        tabs["Transaction"]["Any Level Sectors"] = tabs["Transaction"]["Helper_Any Level Sectors"].apply(replace_sector_names)
        st.write("<small>Updated 'Any Level Sectors' based on 'Helper_Any Level Sectors'</small>", unsafe_allow_html=True)

      
        # Process 'Transaction' tab specific replacements
        process_transaction_sheet(tabs["Transaction"])
        

        # Populating 'Events' tab
        event_columns = [
            ("Current status date", "Current status"),
            ("Financial close", "Financial Close"),
            ("Transaction Launch", "Announced"),
            ("RFP returned", "Request for Proposals"),
            ("Preferred Proponents", "Preferred Bidder"),
            ("Expressions of Interest", "Expression of Interest"),
            ("RFQ returned", "Request for Qualifications"),
            ("Shortlisted proponents", "Shortlist")
        ]

        for date_col, event_type in event_columns:
            if (date_col in df.columns) and (df[date_col].notna().any()):
                event_data = df[[date_col, 'Transaction Upload ID']].dropna(subset=[date_col])
                event_data = event_data[event_data[date_col].apply(lambda x: x != "N/A")]
                event_data = event_data.rename(columns={date_col: 'Event Date'})
                if event_type == "Current status":
                    event_data['Event Type'] = df['Current status']
                else:
                    event_data['Event Type'] = event_type
                tabs["Events"] = pd.concat([tabs["Events"], event_data], ignore_index=True)
                st.write(f"<small>Copied data for 'Events' from '{date_col}'</small>", unsafe_allow_html=True)
            else:
                st.warning(f"'{date_col}' column not found or contains no data in the source file.")

        # Apply replacement for 'Event Type'
        tabs["Events"]['Event Type'] = tabs["Events"]['Event Type'].apply(replace_event_type)

        # Ensure 'Event Date' column is treated as a string before performing string operations
        tabs["Events"]['Event Date'] = tabs["Events"]['Event Date'].astype(str)

        # Remove rows where 'Event Date' is blank/whitespace/na/nan        
        tabs["Events"]['Event Date'] = pd.to_datetime(tabs["Events"]['Event Date'], errors='coerce').dt.normalize()

        # Ensure 'Event Date' column only contains the date, not the time
        tabs["Events"]['Event Date'] = pd.to_datetime(tabs["Events"]['Event Date'], errors='coerce').dt.date

        # Remove duplicate rows
        tabs["Events"] = tabs["Events"].drop_duplicates()



        # Populating 'Bidders_Any' tab
        sources = {
            "Legal Advisors": "Legal Adviser",
            "Technical Advisors": "Technical Adviser",
            "Financial Advisors": "Financial Adviser",
            "Vendors": "Divestor",
            "Grantors": "Awarding Authority"
        }

        entries = []
        for source_column, role_type in sources.items():
            if source_column in df.columns:
                for _, row in df.iterrows():
                    cell_value = row[source_column]
                    if pd.notna(cell_value):
                        cell_value = str(cell_value)  # Ensure the value is a string before splitting

                        # Determine the delimiter based on the column name
                        delimiter = ',' if source_column in ["Vendors", "Grantors"] else ';'
                        
                        companies = cell_value.split(delimiter)  # Split by the determined delimiter
                        for company in companies:
                            company = company.strip()
                            if company:  # Ensure company is not empty
                                # Determine Client Counterparty based on the content in parentheses
                                client_counterparty = ''
                                if '(Funders)' in company:
                                    client_counterparty = 'Debt Provider'
                                elif '(Acquirer)' in company or '(Acquiror)' in company:
                                    client_counterparty = 'Acquirer'
                                elif '(SPV)' in company:
                                    client_counterparty = 'SPV'
                                elif '(Seller)' in company:
                                    client_counterparty = 'Divestor'
                                elif '(Grantor)' in company:
                                    client_counterparty = 'Awarding Authority'
                                elif '(Target)' in company or '(Target Company)' in company:
                                    client_counterparty = 'Target'
                                elif '(Lenders)' in company:
                                    client_counterparty = 'Debt Provider'

                                # Remove parentheses and their content from the company name
                                company_cleaned = re.sub(r'\s*\(.*?\)\s*', '', company).strip()

                                entries.append({
                                    "Transaction Upload ID": row["Transaction Upload ID"],
                                    "Role Type": role_type,
                                    "Role Subtype": "",
                                    "Company": company_cleaned,
                                    "Fund": "",
                                    "Bidder Status": "Successful",
                                    "Client Counterparty": client_counterparty,
                                    "Client Company Name": "",
                                    "Fund Name": ""
                                })
                st.write(f"<small>Copied data for 'Bidders_Any' from '{source_column}'</small>", unsafe_allow_html=True)
            else:
                st.warning(f"'{source_column}' column not found in the source file.")

        if entries:
            tabs["Bidders_Any"] = pd.DataFrame(entries)
        
        # Populating 'Tranches' tab
        tranche_entries = []

        # Ensure the 'Tranches' tab contains all required columns in the specified order
        required_columns = [
            "Transaction Upload ID", "Tranche Upload ID", "Tranche Primary Type", 
            "Tranche Secondary Type", "Tranche Tertiary Type", "Value", 
            "Maturity Start Date", "Maturity End Date", "Tenor", 
            "Tranche ESG Type", "Helper_Tranche Value USD m", "Helper_Transaction Value USD m", "Helper_Transaction Value LC",
            "Helper_Tranche Value USD m as % of Helper_Transaction Value USD m"
        ]

        # Loan Debt Tranche data
        for i in range(1, 21):
            tranche_col = f"Loan Debt Tranche {i} Type"
            tranche_volume_col = f"Tranche {i} Volume USD (m)"
            copied_tranche_col = False  # Flag to track if message is printed for this column
            if tranche_col in df.columns or tranche_volume_col in df.columns:
                # Ensure 'Transaction Upload ID' column exists
                if 'Transaction Upload ID' in df.columns:
                    valid_entries = df[['Transaction Upload ID', tranche_col, tranche_volume_col]].dropna(subset=[tranche_col, tranche_volume_col]).copy()
                    valid_entries['Tranche Upload ID'] = valid_entries['Transaction Upload ID'].astype(str) + f"-L{i}"
                    valid_entries['Tranche Tertiary Type'] = valid_entries[tranche_col]
                    valid_entries['Helper_Tranche Value USD m'] = valid_entries[tranche_volume_col]
                    valid_entries['Helper_Transaction Value USD m'] = df['Transaction size USD(m)'] if 'Transaction size USD(m)' in df.columns else None
                    valid_entries['Helper_Transaction Value LC'] = df['Transaction size (m)'] if 'Transaction size (m)' in df.columns else None
                    valid_entries['Tenor'] = df[f'Tranche {i} Tenor'] if f'Tranche {i} Tenor' in df.columns else None

                    # Remove rows with empty or whitespace 'Tranche Tertiary Type' values
                    valid_entries = valid_entries[valid_entries['Tranche Tertiary Type'].str.strip().astype(bool)]

                    tranche_entries.append(valid_entries)
                    if not copied_tranche_col:
                        st.write(f"<small>Copied data for 'Tranches' from '{tranche_col}' and '{tranche_volume_col}'</small>", unsafe_allow_html=True)
                        copied_tranche_col = True  # Set the flag to True after printing the message
                else:
                    st.warning(f"'Transaction Upload ID' column not found in the source file.")
            else:
                st.warning(f"'{tranche_col}' or '{tranche_volume_col}' column not found in the source file.")

        # Capital Market Debt data
        for i in range(1, 21):
            cap_market_debt_column = f'Capital Market Debt {i} Volume USD (m)'
            copied_cap_market_debt_column = False  # Flag to track if message is printed for this column
            if cap_market_debt_column in df.columns:
                for _, row in df.iterrows():
                    if pd.notna(row[cap_market_debt_column]):
                        volume_usd = row[cap_market_debt_column]
                        transaction_upload_id = row["Transaction Upload ID"]
                        tranche_upload_id = f'{transaction_upload_id}-CM{i}'

                        # Create a temporary DataFrame for the new tranche
                        temp_df = pd.DataFrame({
                            "Transaction Upload ID": [transaction_upload_id],
                            "Tranche Upload ID": [tranche_upload_id],
                            "Tranche Primary Type": [""],
                            "Tranche Secondary Type": [""],
                            "Tranche Tertiary Type": [""],
                            "Value": [""],
                            "Maturity Start Date": [""],
                            "Maturity End Date": [""],
                            "Tenor": [""],
                            "Tranche ESG Type": [""],
                            "Helper_Tranche Value USD m": [volume_usd],
                            "Helper_Transaction Value USD m": [row.get("Transaction size USD(m)", "")],
                            "Helper_Transaction Value LC": [row.get("Transaction size (m)", "")]
                        })

                        # Append the temporary DataFrame to the main tranches_df DataFrame
                        tranche_entries.append(temp_df)
                        if not copied_cap_market_debt_column:
                            st.write(f"<small>Copied data for 'Tranches' from '{cap_market_debt_column}'</small>", unsafe_allow_html=True)
                            copied_cap_market_debt_column = True  # Set the flag to True after printing the message
            else:
                st.warning(f"'{cap_market_debt_column}' column not found in the source file.")

        # Append additional data based on 'Equity Providers at FC'
        if 'Equity Providers at FC' in df.columns:
            equity_providers_df = df.dropna(subset=['Equity Providers at FC'])
            for _, row in equity_providers_df.iterrows():
                transaction_upload_id = row["Transaction Upload ID"]
                tranche_upload_id = f'{transaction_upload_id}-E'
                equity_value = row.get('Equity at FC USD(m)', '')

                temp_df = pd.DataFrame({
                    "Transaction Upload ID": [transaction_upload_id],
                    "Tranche Upload ID": [tranche_upload_id],
                    "Tranche Primary Type": [""],
                    "Tranche Secondary Type": [""],
                    "Tranche Tertiary Type": ["Equity"],
                    "Value": [""],
                    "Maturity Start Date": [""],
                    "Maturity End Date": [""],
                    "Tenor": [""],
                    "Tranche ESG Type": [""],
                    "Helper_Tranche Value USD m": [equity_value],
                    "Helper_Transaction Value USD m": [row.get("Transaction size USD(m)", "")],
                    "Helper_Transaction Value LC": [row.get("Transaction size (m)", "")]
                })
                
                tranche_entries.append(temp_df)
            st.write("<small>Copied data for 'Tranches' from 'Equity Providers at FC'</small>", unsafe_allow_html=True)  # Progress message

        # Concatenate all entries into the 'Tranches' tab if there are any entries
        if tranche_entries:
            tabs["Tranches"] = pd.concat(tranche_entries, ignore_index=True)

        # Update 'Tranche Primary Type', 'Tranche Secondary Type' and 'Tranche Tertiary Type' based on 'Tranche Upload ID'
        tabs["Tranches"]['Tranche Primary Type'] = tabs["Tranches"]['Tranche Upload ID'].apply(
            lambda x: 'Debt' if any(x.endswith(suffix) for suffix in ['L1', 'L2', 'L3', 'CM1', 'CM2', 'CM3']) else 'Equity'
        )
        tabs["Tranches"]['Tranche Secondary Type'] = tabs["Tranches"]['Tranche Upload ID'].apply(
            lambda x: 'Loan' if any(x.endswith(suffix) for suffix in ['L1', 'L2', 'L3']) else ('Bond' if any(x.endswith(suffix) for suffix in ['CM1', 'CM2', 'CM3']) else 'Equity')
        )
        tabs["Tranches"]['Tranche Tertiary Type'] = tabs["Tranches"].apply(
            lambda row: 'Commercial Bond' if any(row['Tranche Upload ID'].endswith(suffix) for suffix in ['CM1', 'CM2', 'CM3']) else row['Tranche Tertiary Type'],
            axis=1
        )

        # Add the new column 'Helper_Tranche Value USD m as % of Helper_Transaction Value USD m'
        tabs["Tranches"]["Helper_Tranche Value USD m as % of Helper_Transaction Value USD m"] = tabs["Tranches"].apply(
            lambda row: safe_float_conversion(row["Helper_Tranche Value USD m"]) / safe_float_conversion(row["Helper_Transaction Value USD m"]) 
            if safe_float_conversion(row["Helper_Transaction Value USD m"]) != 0 else 0, axis=1
        )

        # Populate column F "Value" with results of multiplying columns "Helper_Tranche Value USD m as % of Helper_Transaction Value USD m" by "Helper_Transaction Value LC"
        tabs["Tranches"]["Value"] = tabs["Tranches"].apply(
            lambda row: safe_float_conversion(row["Helper_Tranche Value USD m as % of Helper_Transaction Value USD m"]) * safe_float_conversion(row["Helper_Transaction Value LC"]) 
            if safe_float_conversion(row["Helper_Tranche Value USD m as % of Helper_Transaction Value USD m"]) and safe_float_conversion(row["Helper_Transaction Value LC"]) else 0, axis=1
        )
        
        # Update 'Tranche ESG Type' if 'Tranche Tertiary Type' contains 'Islamic'
        tabs["Tranches"]["Tranche ESG Type"] = tabs["Tranches"].apply(
            lambda row: f'{row["Tranche ESG Type"]}, Tranche ESG Type' if "Islamic" in row["Tranche Tertiary Type"] else row["Tranche ESG Type"],
            axis=1
        )

        # Ensure all required columns are present in the 'Tranches' tab
        for col in required_columns:
            if col not in tabs["Tranches"].columns:
                tabs["Tranches"][col] = None

        tabs["Tranches"] = tabs["Tranches"][required_columns]

        # Replace words in 'Tranche Tertiary Type' column
        tabs["Tranches"] = replace_tranche_tertiary_type(tabs["Tranches"])

        ### Function to populate 'Tranche_Roles_Any' tab ###
        def populate_tranche_roles_any(transaction_df, tranche_roles_any_df):
            entries = []

            # Process 'Tranche 1 Lenders' to 'Tranche 20 Lenders' first
            for i in range(1, 21):
                lenders_column = f'Tranche {i} Lenders'

                if lenders_column in transaction_df.columns:
                    for _, row in transaction_df.iterrows():
                        if pd.notna(row[lenders_column]):
                            lenders = re.split(r',\s*(?![^()]*\))', row[lenders_column])  # Split by comma unless within parentheses
                            for lender in lenders:
                                lender = lender.strip()
                                percentage = ''
                                match = re.search(r'(\d+(\.\d+)?)%\)', lender)  # Adjusted regex to capture percentages
                                if match:
                                    percentage = match.group(1)
                                if lender:
                                    entries.append({
                                        "Transaction Upload ID": row["Transaction Upload ID"],
                                        "Tranche Upload ID": f'{row["Transaction Upload ID"]}-L{i}',
                                        "Tranche Role Type": "",
                                        "Company": lender,
                                        "Fund": "",
                                        "Value": "",
                                        "Percentage": percentage,  # Store the extracted percentage here
                                        "Comment": ""
                                    })

            # Process 'Capital Market Debt 1 Underwriters' to 'Capital Market Debt 20 Underwriters' next
            for i in range(1, 21):
                cm1_column = f'Capital Market Debt {i} Underwriters'
                cm2_column = f'Capital Market Debt 2{i} Underwriters'

                if cm1_column in transaction_df.columns:
                    for _, row in transaction_df.iterrows():
                        if pd.notna(row[cm1_column]):
                            underwriters_cm1 = re.split(r',\s*(?![^()]*\))', row[cm1_column])  # Split by comma unless within parentheses
                            for underwriter in underwriters_cm1:
                                underwriter = underwriter.strip()
                                percentage = ''
                                match = re.search(r'(\d+(\.\d+)?)%\)', underwriter)  # Adjusted regex to capture percentages
                                if match:
                                    percentage = match.group(1)
                                if underwriter:
                                    entries.append({
                                        "Transaction Upload ID": row["Transaction Upload ID"],
                                        "Tranche Upload ID": f'{row["Transaction Upload ID"]}-CM{i}',
                                        "Tranche Role Type": "",
                                        "Company": underwriter,
                                        "Fund": "",
                                        "Value": "",
                                        "Percentage": percentage,  # Store the extracted percentage here
                                        "Comment": ""
                                    })

                if cm2_column in transaction_df.columns:
                    for _, row in transaction_df.iterrows():
                        if pd.notna(row[cm2_column]):
                            underwriters_cm2 = re.split(r',\s*(?![^()]*\))', row[cm2_column])  # Split by comma unless within parentheses
                            for underwriter in underwriters_cm2:
                                underwriter = underwriter.strip()
                                percentage = ''
                                match = re.search(r'(\d+(\.\d+)?)%\)', underwriter)  # Adjusted regex to capture percentages
                                if match:
                                    percentage = match.group(1)
                                if underwriter:
                                    entries.append({
                                        "Transaction Upload ID": row["Transaction Upload ID"],
                                        "Tranche Upload ID": f'{row["Transaction Upload ID"]}-CM2{i}',
                                        "Tranche Role Type": "",
                                        "Company": underwriter,
                                        "Fund": "",
                                        "Value": "",
                                        "Percentage": percentage,  # Store the extracted percentage here
                                        "Comment": ""
                                    })

            # Append additional data based on 'Equity Providers at FC'
            if 'Equity Providers at FC' in transaction_df.columns:
                equity_providers_df = transaction_df.dropna(subset=['Equity Providers at FC'])
                for _, row in equity_providers_df.iterrows():
                    equity_providers = re.split(r',\s*(?![^()]*\))', row['Equity Providers at FC'])  # Split by comma unless within parentheses
                    for provider in equity_providers:
                        provider = provider.strip()
                        percentage = ''
                        match = re.search(r'(\d+(\.\d+)?)%\)', provider)  # Adjusted regex to capture percentages
                        if match:
                            percentage = match.group(1)
                        if provider:
                            entries.append({
                                "Transaction Upload ID": row["Transaction Upload ID"],
                                "Tranche Upload ID": f'{row["Transaction Upload ID"]}-E',
                                "Tranche Role Type": "",
                                "Company": provider,
                                "Fund": "",
                                "Value": "",
                                "Percentage": percentage,  # Store the extracted percentage here
                                "Comment": ""
                            })

            # Create DataFrame from list of dictionaries
            tranche_roles_any_df = pd.DataFrame(entries, columns=[
                "Transaction Upload ID", "Tranche Upload ID", "Tranche Role Type", "Company", "Fund", 
                "Value", "Percentage", "Comment"
            ])

            return tranche_roles_any_df

        # Populate 'Tranche_Roles_Any' tab
        tabs["Tranche_Roles_Any"] = populate_tranche_roles_any(df, tabs["Tranche_Roles_Any"])
        st.write("<small>Copied data for 'Tranche_Roles_Any'</small>", unsafe_allow_html=True)  # Progress message

        # Clean company names in 'Tranche_Roles_Any' tab
        tabs["Tranche_Roles_Any"] = clean_company_names(tabs["Tranche_Roles_Any"])

        # Additional functionality to update 'Tranche Role Type' in 'Tranche_Roles_Any' tab
        def update_tranche_roles_any(tranches_df, tranche_roles_any_df):
            for _, row in tranches_df.iterrows():
                tranche_upload_id = row['Tranche Upload ID']
                primary_type = row['Tranche Primary Type']
                secondary_type = row['Tranche Secondary Type']

                if primary_type == 'Equity':
                    tranche_roles_any_df.loc[
                        tranche_roles_any_df['Tranche Upload ID'] == tranche_upload_id,
                        'Tranche Role Type'
                    ] = "Sponsor"

                if secondary_type == 'Bond':
                    tranche_roles_any_df.loc[
                        tranche_roles_any_df['Tranche Upload ID'] == tranche_upload_id,
                        'Tranche Role Type'
                    ] = "Bond Arranger"

                if secondary_type == 'Loan' or secondary_type == 'Debt':
                    tranche_roles_any_df.loc[
                        tranche_roles_any_df['Tranche Upload ID'] == tranche_upload_id,
                        'Tranche Role Type'
                    ] = "Debt Provider"

            return tranche_roles_any_df

        # Apply the update to 'Tranche_Roles_Any' tab
        tabs["Tranche_Roles_Any"] = update_tranche_roles_any(tabs["Tranches"], tabs["Tranche_Roles_Any"])

        # Save the DataFrame to a new Excel file with multiple sheets
        london_tz = pytz.timezone('Europe/London')
        current_time = datetime.now(london_tz)
        formatted_time = current_time.strftime('%Y%m%d_%H%M')
        new_file_name = f"curated_INFRA3_{formatted_time}.xlsx"
        new_file_path = os.path.join(tempfile.gettempdir(), new_file_name)
        
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            for tab_name, tab_df in tabs.items():
                tab_df.to_excel(writer, sheet_name=tab_name, index=False)
        
        # Autofit columns
        autofit_columns(new_file_path)
        
        end_time = time.time()
        elapsed_time = end_time - start_time
        elapsed_time_str = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
        st.write(f"<small>Processing completed in {elapsed_time_str}.</small>", unsafe_allow_html=True)
        
        return new_file_path

    except Exception as e:
        end_time = time.time()
        elapsed_time = end_time - start_time
        elapsed_time_str = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
        st.error(f"An error occurred: {e}")
        st.write(f"<small>Processing stopped at {elapsed_time_str} due to the error.</small>", unsafe_allow_html=True)
        raise

### Streamlit app ###
st.title('Curating INFRA 3i Data Files')

uploaded_file = st.file_uploader("Choose a source file", type=["xlsx"])

if uploaded_file is not None:
    start_time = time.time()  # Start the timer once a file is uploaded
    
    # Display the timer on the frontend
    while not st.spinner("Processing the file..."):
        elapsed_time = time.time() - start_time
        elapsed_time_str = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))
        st.write(f"Elapsed time: {elapsed_time_str}")
        time.sleep(1)

    # Save the uploaded file to a temporary directory
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.write(uploaded_file.getbuffer())
    temp_file_path = temp_file.name
    temp_file.close()  # Ensure file is closed before processing

    destination_path = None  # Initialize destination_path

    try:
        with st.spinner("Processing the file..."):
            destination_path = create_destination_file(temp_file_path, start_time)
        st.success("File processed successfully!")

        # Provide a download button for the processed file
        with open(destination_path, "rb") as file:
            st.download_button(
                label="Download Processed File",
                data=file,
                file_name=os.path.basename(destination_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"An error occurred: {e}")

    finally:
        # Clean up temporary files
        try:
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
        except PermissionError:
            st.warning("Temporary file could not be deleted immediately, please try again later.")
        if destination_path and os.path.exists(destination_path):
            os.remove(destination_path)

else:
    st.info("Please upload an Excel file to start processing.")
